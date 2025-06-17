import streamlit as st
import tempfile
import os
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill
from copy import copy

def inject_monthly_sheets_with_formatting(template_path, user_file_path, output_path):
    template_wb = load_workbook(template_path, data_only=False)
    user_wb = load_workbook(user_file_path)
    monthly_sheets = [name for name in template_wb.sheetnames if name.startswith("2025")]

    for sheet in monthly_sheets:
        if sheet in user_wb.sheetnames:
            del user_wb[sheet]

    for sheet_name in monthly_sheets:
        source_ws = template_wb[sheet_name]
        new_ws = user_wb.create_sheet(title=sheet_name)

        for row in source_ws.iter_rows():
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        for row_idx, row_dim in source_ws.row_dimensions.items():
            new_ws.row_dimensions[row_idx].height = row_dim.height
        for col_letter, col_dim in source_ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = col_dim.width
        for merged_range in source_ws.merged_cells.ranges:
            new_ws.merge_cells(str(merged_range))

    user_wb.save(output_path)

def copy_updated_sheets_with_formatting(template_path, target_path, output_path):
    template_wb = load_workbook(template_path, data_only=False)
    target_wb = load_workbook(target_path)

    sheet_mappings = {'v1': 'Γενικό Αποτέλεσμα', 'v2': 'Διαφορά'}

    for source, target in sheet_mappings.items():
        if target in target_wb.sheetnames:
            del target_wb[target]
        source_ws = template_wb[source]
        new_ws = target_wb.create_sheet(title=target)
        for row in source_ws.iter_rows():
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.column)
                if cell.data_type == "f":
                    formula = str(cell.value)
                    if source == "v2":
                        formula = formula.replace("'v1'", "'Γενικό Αποτέλεσμα'")
                    new_cell.value = f"={formula}" if not formula.startswith("=") else formula
                else:
                    new_cell.value = cell.value
                if cell.has_style:
                    new_cell._style = cell._style

    formatting_ws = target_wb['Γενικό Αποτέλεσμα']
    cols = ["C", "E", "G", "I", "K", "M", "O", "Q", "S", "U", "W", "Y"]
    row_ranges = [
        range(36, 61), range(64, 89), range(96, 124), range(128, 158),
        range(162, 202), range(207, 236), range(240, 270), range(273, 313)
    ]
    fill = PatternFill(start_color="6699FF", end_color="6699FF", fill_type="solid")
    for rows in row_ranges:
        for r in rows:
            for col in cols:
                formatting_ws[f"{col}{r}"].fill = fill

    target_wb.save(output_path)

# === Streamlit Greek UI ===
st.set_page_config(page_title="Βοηθός Excel", layout="centered")
st.title("📊 Βοηθός Excel — Εισαγωγή Μηνιαίων & Συνοπτικών Φύλλων")

st.markdown("""
Ανεβάστε το αρχείο Excel με τα βασικά σας φύλλα. Η εφαρμογή θα:
1. Εισάγει τα μηνιαία φύλλα (2025-ΧΧ)
2. Δημιουργήσει τα συνοπτικά φύλλα:
   - `Γενικό Αποτέλεσμα`
   - `Διαφορά`
""")

uploaded_file = st.file_uploader("📁 Μεταφόρτωση αρχείου Excel", type=["xlsx"])

if uploaded_file:
    base_filename = os.path.splitext(uploaded_file.name)[0]
    updated_filename = f"{base_filename}_updated.xlsx"

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_input, \
         tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_after_months, \
         tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_final:

        tmp_input.write(uploaded_file.read())
        tmp_input.flush()

        inject_monthly_sheets_with_formatting(
            template_path="template_phase_2_cleaned.xlsx",
            user_file_path=tmp_input.name,
            output_path=tmp_after_months.name
        )

        copy_updated_sheets_with_formatting(
            template_path="bilio_with_v3_formulas.xlsx",
            target_path=tmp_after_months.name,
            output_path=tmp_final.name
        )

        st.success("✅ Ολοκληρώθηκε! Το αρχείο σας είναι έτοιμο.")

        with open(tmp_final.name, "rb") as f:
            st.download_button(
                label="📥 Λήψη ενημερωμένου αρχείου Excel",
                data=f,
                file_name=updated_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
